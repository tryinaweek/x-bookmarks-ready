import os
import json
import secrets
import hashlib
import base64
import urllib.parse
import io
from datetime import datetime, timezone, timedelta

import requests as req_lib
import anthropic
from flask import Flask, render_template, request, redirect, session, send_file

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import firebase_admin
from firebase_admin import credentials, firestore, auth

app = Flask(__name__, template_folder="../templates", static_folder="../static")
app.secret_key = os.environ.get("SESSION_SECRET", secrets.token_hex(32))

@app.context_processor
def inject_firebase_config():
    api_key = os.environ.get("FIREBASE_API_KEY", "").strip()
    if api_key.startswith('"') and api_key.endswith('"'):
        api_key = api_key[1:-1].strip()
    if api_key.startswith("'") and api_key.endswith("'"):
        api_key = api_key[1:-1].strip()
        
    project_id = os.environ.get("FIREBASE_PROJECT_ID", "x-bookmark-app").strip()
    if project_id.startswith('"') and project_id.endswith('"'):
        project_id = project_id[1:-1].strip()
    if project_id.startswith("'") and project_id.endswith("'"):
        project_id = project_id[1:-1].strip()

    return {
        "firebase_config": {
            "apiKey": api_key,
            "authDomain": f"{project_id}.firebaseapp.com",
            "projectId": project_id,
        }
    }

@app.context_processor
def inject_twitter_status():
    db_uid = session.get("db_user_id") or session.get("firebase_uid")
    twitter_connected = False
    if db_uid:
        try:
            token = get_valid_twitter_token(db_uid)
            twitter_connected = bool(token)
        except Exception:
            pass
    return dict(twitter_connected=twitter_connected)


CLIENT_ID = os.environ.get("X_CLIENT_ID", "")
CLIENT_SECRET = os.environ.get("X_CLIENT_SECRET", "")
CLAUDE_API_KEY = os.environ.get("CLAUDE_API_KEY", "")

# Initialize Firebase Admin & Firestore client defensively
db = None
startup_error = None

try:
    try:
        firebase_admin.get_app()
    except ValueError:
        local_key = os.path.join(os.path.dirname(__file__), "../service-account.json")
        if os.path.exists(local_key):
            cred = credentials.Certificate(local_key)
            firebase_admin.initialize_app(cred)
        else:
            firebase_project_id = os.environ.get("FIREBASE_PROJECT_ID")
            firebase_client_email = os.environ.get("FIREBASE_CLIENT_EMAIL")
            firebase_private_key = os.environ.get("FIREBASE_PRIVATE_KEY")
            
            if firebase_project_id and firebase_client_email and firebase_private_key:
                formatted_key = firebase_private_key.replace("\\n", "\n").strip()
                # Remove surrounding quotes if they were mistakenly copied
                if formatted_key.startswith('"') and formatted_key.endswith('"'):
                    formatted_key = formatted_key[1:-1].strip()
                if formatted_key.startswith("'") and formatted_key.endswith("'"):
                    formatted_key = formatted_key[1:-1].strip()
                cred = credentials.Certificate({
                    "type": "service_account",
                    "project_id": firebase_project_id,
                    "client_email": firebase_client_email,
                    "private_key": formatted_key,
                    "token_uri": "https://oauth2.googleapis.com/token",
                })
                firebase_admin.initialize_app(cred)
            else:
                missing = []
                if not firebase_project_id: missing.append("FIREBASE_PROJECT_ID")
                if not firebase_client_email: missing.append("FIREBASE_CLIENT_EMAIL")
                if not firebase_private_key: missing.append("FIREBASE_PRIVATE_KEY")
                raise ValueError(f"Missing required environment variables for Firebase initialization: {', '.join(missing)}")
    db = firestore.client()
except Exception as e:
    import traceback
    startup_error = {
        "error": str(e),
        "traceback": traceback.format_exc()
    }


@app.before_request
def check_startup_error():
    if startup_error:
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>MyBookmarks - Configuration Error</title>
            <style>
                body {{ font-family: -apple-system, system-ui; padding: 40px; background: #fef2f2; color: #991b1b; line-height: 1.6; }}
                .card {{ background: white; border: 1px solid #fee2e2; border-radius: 12px; padding: 32px; max-width: 800px; margin: 0 auto; box-shadow: 0 4px 6px rgba(0,0,0,0.05); }}
                h1 {{ font-size: 24px; font-weight: 700; margin-top: 0; display: flex; align-items: center; gap: 8px; }}
                pre {{ background: #f8fafc; border: 1px solid #e2e8f0; color: #334155; padding: 16px; border-radius: 8px; font-family: monospace; font-size: 13px; overflow-x: auto; white-space: pre-wrap; }}
                .badge {{ background: #fee2e2; color: #991b1b; font-size: 11px; padding: 4px 8px; border-radius: 4px; font-weight: 600; text-transform: uppercase; }}
            </style>
        </head>
        <body>
            <div class="card">
                <h1><span class="badge">Configuration Error</span> Firebase Initialization Failed</h1>
                <p>The server failed to connect to Firebase. This usually happens when environment variables are missing or formatted incorrectly in your Vercel project settings.</p>
                <h3>Error Message:</h3>
                <pre>{startup_error['error']}</pre>
                <h3>Traceback Details:</h3>
                <pre>{startup_error['traceback']}</pre>
                <p style="color: #475569; font-size: 13px; margin-top: 24px;">Please check your Vercel environment variables and trigger a redeployment once fixed.</p>
            </div>
        </body>
        </html>
        """, 500

OWNER_X_ID = os.environ.get("OWNER_X_ID", "25914613")  # Your X user ID - full access

AUTH_URL = "https://twitter.com/i/oauth2/authorize"
TOKEN_URL = "https://api.twitter.com/2/oauth2/token"
SCOPES = "bookmark.read tweet.read tweet.write users.read offline.access"


# -- DB helpers (Cloud Firestore) ----------------------------------------------

def get_valid_twitter_token(user_id):
    """Checks and retrieves a valid Twitter access token, refreshing it if expired."""
    if not user_id:
        return None
    user_ref = db.collection("users").document(user_id)
    user_doc = user_ref.get()
    if not user_doc.exists:
        return None
    user_data = user_doc.to_dict()
    
    access_token = user_data.get("access_token")
    refresh_token = user_data.get("refresh_token")
    expires_at = user_data.get("expires_at", 0)
    
    # If not expired, return access token
    if access_token and datetime.now(timezone.utc).timestamp() < expires_at:
        return access_token
        
    # If expired and we have a refresh token, refresh it
    if refresh_token:
        try:
            r = req_lib.post(TOKEN_URL, auth=(CLIENT_ID, CLIENT_SECRET),
                              data={"grant_type": "refresh_token", "refresh_token": refresh_token})
            data = r.json()
            if "access_token" in data:
                new_access = data["access_token"]
                new_refresh = data.get("refresh_token", refresh_token)
                new_expires_at = datetime.now(timezone.utc).timestamp() + data.get("expires_in", 7200) - 60
                
                user_ref.update({
                    "access_token": new_access,
                    "refresh_token": new_refresh,
                    "expires_at": new_expires_at,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                })
                return new_access
        except Exception as e:
            print(f"Token refresh failed for user {user_id}: {e}")
            
    return access_token


def db_get_or_create_user(x_user_id, username, name, access_token, refresh_token=None):
    firebase_uid = session.get("firebase_uid")
    if not firebase_uid:
        # Fallback to finding user by x_user_id
        docs = db.collection("users").where("x_user_id", "==", str(x_user_id)).limit(1).get()
        if docs:
            firebase_uid = docs[0].id
        else:
            firebase_uid = f"x_{x_user_id}"
            
    user_ref = db.collection("users").document(firebase_uid)
    doc = user_ref.get()
    
    data = {
        "x_user_id": str(x_user_id),
        "username": username,
        "name": name,
        "access_token": access_token,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    if refresh_token:
        data["refresh_token"] = refresh_token
        
    if doc.exists:
        user_ref.update(data)
    else:
        data["created_at"] = datetime.now(timezone.utc).isoformat()
        user_ref.set(data)
    return firebase_uid


def db_save_cache(table, user_id, data, last_id=None):
    if not user_id:
        return
    cache_ref = db.collection("users").document(user_id).collection("caches").document(table)
    row = {
        "data": data,
        "fetched_at": datetime.now(timezone.utc).isoformat()
    }
    if last_id:
        row["last_id"] = last_id
    cache_ref.set(row)


def db_load_cache_full(table, user_id):
    if not user_id:
        return None, None, None
    doc = db.collection("users").document(user_id).collection("caches").document(table).get()
    if doc.exists:
        d = doc.to_dict()
        return d.get("data"), d.get("last_id"), d.get("fetched_at")
    return None, None, None


def db_merge_cache(table, user_id, new_data, last_id):
    existing, _, _ = db_load_cache_full(table, user_id)
    if existing:
        existing_ids = {item["id"] for item in existing}
        merged = new_data + [item for item in existing if item["id"] not in existing_ids]
    else:
        merged = new_data
    db_save_cache(table, user_id, merged, last_id)
    return merged


def db_save_analysis(user_id, analysis_type, data):
    if not user_id:
        return
    ref = db.collection("users").document(user_id).collection("analyses").document(analysis_type)
    ref.set({
        "data": data,
        "created_at": datetime.now(timezone.utc).isoformat()
    })


def db_load_analysis(user_id, analysis_type):
    if not user_id:
        return None
    doc = db.collection("users").document(user_id).collection("analyses").document(analysis_type).get()
    if doc.exists:
        return doc.to_dict().get("data")
    return None


def db_save_suggestions(user_id, data):
    if not user_id:
        return
    ref = db.collection("users").document(user_id).collection("suggestions").document("current")
    ref.set({
        "data": data,
        "generated_at": datetime.now(timezone.utc).isoformat()
    })


def db_load_suggestions(user_id):
    if not user_id:
        return None
    doc = db.collection("users").document(user_id).collection("suggestions").document("current").get()
    if doc.exists:
        return doc.to_dict().get("data")
    return None


def db_save_profile(user_id, data):
    if not user_id:
        return
    user_ref = db.collection("users").document(user_id)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    user_ref.update(data)


def db_load_profile(user_id):
    if not user_id:
        return None
    doc = db.collection("users").document(user_id).get()
    return doc.to_dict() if doc.exists else None


def get_voice_context(user_id):
    """Build a rich context string from the user's profile for AI prompts."""
    profile = _safe_db(db_load_profile, user_id)
    if not profile:
        return PROFILE_CONTEXT

    parts = []
    if profile.get("bio"):
        parts.append(f"Bio: {profile['bio']}")
    if profile.get("expertise"):
        parts.append(f"Expertise: {profile['expertise']}")
    if profile.get("current_focus"):
        parts.append(f"Current focus: {profile['current_focus']}")
    if profile.get("opinions"):
        parts.append(f"Strong opinions/takes: {profile['opinions']}")
    if profile.get("donts"):
        parts.append(f"NEVER do these in content: {profile['donts']}")

    voice_examples = profile.get("voice_examples", [])
    if voice_examples:
        examples_text = "\n".join(f"- {ex}" for ex in voice_examples if ex)
        if examples_text:
            parts.append(f"Voice examples (write like these):\n{examples_text}")

    if not parts:
        return PROFILE_CONTEXT

    return "\n".join(parts)


def db_save_draft(user_id, tweets, fmt, topic, status="draft"):
    if not user_id:
        return None
    draft_ref = db.collection("users").document(user_id).collection("drafts").document()
    draft_id = draft_ref.id
    draft_ref.set({
        "id": draft_id,
        "tweets": tweets,
        "format": fmt,
        "topic": topic,
        "status": status,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    return draft_id


def db_load_drafts(user_id, status=None):
    if not user_id:
        return []
    # Avoid composite index requirements by sorting/filtering in-memory
    try:
        docs = db.collection("users").document(user_id).collection("drafts").get()
        drafts = []
        for doc in docs:
            d = doc.to_dict()
            # Ensure doc ID is included
            d["id"] = doc.id
            if status and d.get("status") != status:
                continue
            drafts.append(d)
        # Sort in memory descending by created_at
        drafts = sorted(drafts, key=lambda x: x.get("created_at", ""), reverse=True)
        return drafts
    except Exception as e:
        print(f"Error in db_load_drafts: {e}")
        return []


def db_delete_draft(draft_id, user_id):
    if not user_id or not draft_id:
        return
    db.collection("users").document(user_id).collection("drafts").document(draft_id).delete()


def db_mark_posted(draft_id, user_id):
    if not user_id or not draft_id:
        return
    db.collection("users").document(user_id).collection("drafts").document(draft_id).update({
        "status": "posted",
        "posted_at": datetime.now(timezone.utc).isoformat()
    })


# -- X API helpers ---------------------------------------------------------

def get_redirect_uri():
    scheme = request.headers.get("X-Forwarded-Proto", request.scheme)
    host = request.headers.get("X-Forwarded-Host", request.host)
    return f"{scheme}://{host}/callback"


def generate_pkce():
    verifier = secrets.token_urlsafe(32)
    challenge = base64.urlsafe_b64encode(hashlib.sha256(verifier.encode()).digest()).rstrip(b"=").decode()
    return verifier, challenge


def exchange_code(code, verifier):
    r = req_lib.post(TOKEN_URL, auth=(CLIENT_ID, CLIENT_SECRET),
                      data={"grant_type": "authorization_code", "code": code,
                            "redirect_uri": get_redirect_uri(), "code_verifier": verifier})
    return r.json()


def get_me(token):
    r = req_lib.get("https://api.twitter.com/2/users/me", headers={"Authorization": f"Bearer {token}"})
    d = r.json()
    return d["data"]["id"], d["data"]["username"], d["data"]["name"]


def is_owner():
    return str(session.get("firebase_uid", "")) == str(OWNER_X_ID) or str(session.get("user_id", "")) == str(OWNER_X_ID)


def fetch_bookmarks_delta(token, user_id, since_id=None, max_items=None):
    """Fetch bookmarks. If since_id provided, only fetches newer ones."""
    bookmarks, cursor, api_error = [], None, None
    fetch_size = min(max_items, 100) if max_items else 100
    while True:
        params = {"max_results": fetch_size, "tweet.fields": "created_at,text,author_id,public_metrics",
                  "user.fields": "name,username", "expansions": "author_id"}
        if cursor:
            params["pagination_token"] = cursor
        if since_id:
            params["since_id"] = since_id
        r = req_lib.get(f"https://api.twitter.com/2/users/{user_id}/bookmarks",
                         headers={"Authorization": f"Bearer {token}"}, params=params)
        data = r.json()
        if "data" not in data:
            if not bookmarks:
                api_error = data
            break
        users = {u["id"]: u for u in data.get("includes", {}).get("users", [])}
        for t in data["data"]:
            au = users.get(t.get("author_id", ""), {})
            m = t.get("public_metrics", {})
            bookmarks.append({
                "id": t.get("id", ""), "text": t.get("text", ""),
                "name": au.get("name", ""), "username": au.get("username", ""),
                "date": t.get("created_at", "")[:10],
                "likes": m.get("like_count", 0), "retweets": m.get("retweet_count", 0),
                "url": f"https://twitter.com/{au.get('username', '')}/status/{t.get('id', '')}",
            })
        if max_items and len(bookmarks) >= max_items:
            break
        cursor = data.get("meta", {}).get("next_token")
        if not cursor:
            break
    return bookmarks[:max_items] if max_items else bookmarks, api_error


def fetch_tweets_delta(token, user_id, since_id=None, max_pages=5, max_items=None):
    """Fetch user tweets. If since_id provided, only fetches newer ones."""
    tweets, cursor, api_error = [], None, None
    fetch_size = min(max_items, 100) if max_items else 100
    pages = 0
    while pages < max_pages:
        params = {"max_results": fetch_size, "tweet.fields": "created_at,text,public_metrics,source",
                  "exclude": "retweets,replies"}
        if cursor:
            params["pagination_token"] = cursor
        if since_id:
            params["since_id"] = since_id
        r = req_lib.get(f"https://api.twitter.com/2/users/{user_id}/tweets",
                         headers={"Authorization": f"Bearer {token}"}, params=params)
        data = r.json()
        if "data" not in data:
            if not tweets:
                api_error = data
            break
        for t in data["data"]:
            m = t.get("public_metrics", {})
            tweets.append({
                "id": t.get("id", ""), "text": t.get("text", ""),
                "date": t.get("created_at", "")[:10],
                "likes": m.get("like_count", 0), "retweets": m.get("retweet_count", 0),
                "replies": m.get("reply_count", 0), "impressions": m.get("impression_count", 0),
                "url": f"https://twitter.com/i/status/{t.get('id', '')}",
            })
        if max_items and len(tweets) >= max_items:
            break
        cursor = data.get("meta", {}).get("next_token")
        if not cursor:
            break
        pages += 1
    return tweets[:max_items] if max_items else tweets, api_error


# -- AI helpers ------------------------------------------------------------

PROFILE_CONTEXT = """Profile: Angel investor, operator, and tech founder.
Focus areas: AI agents, Claude Code, no-code/low-code, SaaS, startups, entrepreneurship, building in public.
Voice: Direct, practical, experience-driven. Shares frameworks, lessons learned, and actionable insights.
Audience: Founders, builders, developers, AI enthusiasts, indie hackers."""

FORMATTING_RULES = """STRICT FORMATTING RULES:
- SHORT. PUNCHY. Every sentence earns its place. Cut the fluff ruthlessly.
- Contrarian opening that challenges conventional wisdom. Start with a bold claim or hot take.
- Personal proof: back it up with YOUR real experience, specific numbers, actual projects.
- One-line closer that hits hard. Not a paragraph. ONE line.
- Lists: MAX 3-5 items. Never 7+. Each item is one tight line, not a paragraph.
- Line breaks between every thought. White space is your weapon.
- NO generic advice. NO "here's what I learned" followed by obvious points.
- Every post must have a SPECIFIC detail that only YOU would know (project names, dollar amounts, real outcomes).
- Structure: Hook (1 line) -> Proof (2-3 lines) -> Insight (1-2 lines) -> Closer (1 line)
- Threads: First tweet = pure hook, no number. Keep to 4-6 tweets max. Each tweet is TIGHT.
- Under 280 chars per tweet. Shorter is better. The best tweets are under 200 chars.
- End with a sharp question or a bold one-liner. NOT "Bookmark this" or "RT to help others."
- Tone: confident, direct, from experience. Not lecturing. Sharing."""


def _call_claude(prompt, max_tokens=4096):
    client = anthropic.Anthropic(api_key=CLAUDE_API_KEY)
    message = client.messages.create(
        model="claude-3-5-sonnet-latest", max_tokens=max_tokens,
        messages=[{"role": "user", "content": prompt}],
    )
    try:
        raw = message.content[0].text
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[1].rsplit("```", 1)[0]
        return json.loads(raw), None
    except (json.JSONDecodeError, IndexError, KeyError) as e:
        return None, f"Failed to parse AI response: {e}"


def analyze_bookmarks(bookmarks, username=""):
    if not CLAUDE_API_KEY:
        return None, "CLAUDE_API_KEY not configured"
    condensed = [f"[{i}] ({bm['date']}) @{bm['username']}: {bm['text'][:280]}" for i, bm in enumerate(bookmarks, 1)]
    user_ref = f"@{username}" if username else "user"
    prompt = f"""Analyze these {len(bookmarks)} X/Twitter bookmarks for {user_ref}. Return ONLY valid JSON:
{{"summary":"2-3 sentences using you/your","categories":[{{"name":"...","count":5,"bookmark_ids":[1,5],"summary":"..."}}],"timeline":[{{"period":"...","theme":"...","count":15,"bookmark_ids":[1,2]}}],"gems":[{{"id":5,"title":"...","reason":"..."}}],"stale":[{{"id":12,"title":"...","reason":"..."}}],"actions":[{{"text":"...","bookmark_ids":[20,50]}}]}}
Rules: 5-8 categories, 3-5 timeline phases with count, 5-10 gems, stale items, 3-5 actions with bookmark_ids. Use you/your.
Bookmarks:\n""" + "\n".join(condensed)
    return _call_claude(prompt)


def analyze_tweets(tweets, username=""):
    if not CLAUDE_API_KEY:
        return None, "CLAUDE_API_KEY not configured"
    condensed = [f"[{i}] ({tw['date']}) {tw['text'][:280]} | L:{tw['likes']} RT:{tw['retweets']} R:{tw['replies']} I:{tw['impressions']}" for i, tw in enumerate(tweets, 1)]
    prompt = f"""Analyze these {len(tweets)} tweets from @{username}. Return ONLY valid JSON:
{{"summary":"...","top_performers":[{{"id":1,"title":"...","why":"..."}}],"underperformers":[{{"id":5,"title":"...","why":"..."}}],"patterns":[{{"pattern":"...","evidence":"...","recommendation":"..."}}],"content_suggestions":[{{"tweet":"...","based_on":[1],"rationale":"..."}}],"strategy":{{"best_topics":["..."],"avoid_topics":["..."],"best_formats":["..."],"posting_advice":"..."}}}}
Rules: 5-8 top, 3-5 under, 3-5 patterns, 5-8 suggestions under 280 chars, strategy. Use you/your.
Tweets:\n""" + "\n".join(condensed)
    return _call_claude(prompt)


def generate_smart_suggestions(username, db_uid):
    """Generate suggestions using bookmark interests + tweet patterns + trends."""
    if not CLAUDE_API_KEY:
        return []

    # Gather intelligence
    bookmark_analysis = db_load_analysis(db_uid, "bookmarks")
    tweet_analysis = db_load_analysis(db_uid, "tweets")

    voice = get_voice_context(db_uid)
    context_parts = [voice]

    if bookmark_analysis:
        cats = bookmark_analysis.get("categories", [])
        if cats:
            topics = ", ".join(c["name"] for c in cats[:5])
            context_parts.append(f"Current bookmark interests: {topics}")

    if tweet_analysis:
        strategy = tweet_analysis.get("strategy", {})
        if strategy.get("best_topics"):
            context_parts.append(f"Top performing tweet topics: {', '.join(strategy['best_topics'])}")
        if strategy.get("best_formats"):
            fmts = strategy["best_formats"]
            if isinstance(fmts, list):
                context_parts.append(f"Best tweet formats: {', '.join(fmts)}")
            else:
                context_parts.append(f"Best tweet formats: {fmts}")
        patterns = tweet_analysis.get("patterns", [])
        if patterns:
            context_parts.append(f"Proven patterns: {patterns[0].get('pattern', '')}")

    full_context = "\n".join(context_parts)

    prompt = f"""Twitter/X content strategist for @{username}.

{full_context}

Generate 8 tweet/thread ideas that would perform well RIGHT NOW (March 2026).
These ideas should combine:
1. The user's current interests (from their bookmarks)
2. What works for their audience (from their tweet performance)
3. Trending topics in AI, tech, startups this week

Return ONLY valid JSON:
{{"suggestions":[{{"topic":"5-8 words","hook":"Opening line that stops the scroll","format":"tweet or thread","why":"Why this would work based on their data + trends"}}]}}

Rules:
- 4 timely/trending ideas aligned with their interests
- 4 evergreen ideas based on proven engagement patterns
- Hooks in Justin Welsh / Sahil Bloom style
- Each idea should feel personalized, not generic"""

    try:
        result, _ = _call_claude(prompt, max_tokens=2048)
        return result.get("suggestions", []) if result else []
    except Exception:
        return []


LINKEDIN_FORMATTING = """LINKEDIN FORMATTING:
- Contrarian hook that stops the scroll. One bold sentence. Then a blank line.
- Short paragraphs: 1-2 sentences MAX each, with blank lines between.
- Structure: Bold claim -> Personal proof (real projects, real numbers) -> The insight -> Sharp closer.
- Lists: 3-5 items max. Each item is one tight line.
- Include SPECIFIC details: project names, dollar amounts, timeframes, real outcomes.
- NO generic wisdom. Every point must be backed by personal experience.
- End with a sharp one-liner or a direct question. Not "Agree? Drop a comment."
- 3-5 relevant hashtags on the last line.
- Total: 600-1000 characters. Tighter is better.
- Tone: confident operator sharing real experience. Not lecturing."""


def generate_draft(username, idea, format_type, platform="x", db_uid=None):
    if not CLAUDE_API_KEY:
        return [], None
    voice = get_voice_context(db_uid) if db_uid else PROFILE_CONTEXT
    if platform == "linkedin":
        prompt = f"""Create a viral LinkedIn post for {username} about: {idea}

{voice}

IMPORTANT: Write in EXACTLY this person's voice and style. Use their actual perspective, not generic advice.

Return ONLY valid JSON: {{"linkedin_post":"The full post text"}}
{LINKEDIN_FORMATTING}"""
        try:
            result, _ = _call_claude(prompt, max_tokens=2048)
            post = result.get("linkedin_post", "") if result else ""
            return [post] if post else [], "linkedin"
        except Exception:
            return [], "linkedin"
    elif platform == "both":
        prompt = f"""Create content for BOTH Twitter/X and LinkedIn about: {idea}

{voice}

IMPORTANT: Write in EXACTLY this person's voice and style. Use their actual perspective, not generic advice.

Return ONLY valid JSON:
{{"tweets":["Hook tweet (no number)","2/ Second","3/ Third",...],"linkedin_post":"The full LinkedIn post"}}

For Twitter/X thread:
{FORMATTING_RULES}
Write 5-8 tweets. First = pure hook. Last = CTA. Each under 280 chars.

For LinkedIn:
{LINKEDIN_FORMATTING}
Adapt the same core idea but in LinkedIn's longer, storytelling format."""
        try:
            result, _ = _call_claude(prompt, max_tokens=4096)
            if not result:
                return [], "both"
            tweets = result.get("tweets", [])
            li_post = result.get("linkedin_post", "")
            return tweets, "both", li_post
        except Exception:
            return [], "both", ""
    elif format_type == "thread":
        prompt = f"""Create viral Twitter/X thread for @{username} about: {idea}

{voice}

IMPORTANT: Write in EXACTLY this person's voice and style. Use their actual perspective, not generic advice.

Return ONLY valid JSON: {{"tweets":["Hook tweet (no number)","2/ Second","3/ Third",...]}}
{FORMATTING_RULES}
Write 5-8 tweets. First = pure hook. Last = CTA. Each under 280 chars."""
    else:
        prompt = f"""Create single viral tweet for @{username} about: {idea}

{voice}

IMPORTANT: Write in EXACTLY this person's voice and style. Use their actual perspective, not generic advice.

Return ONLY valid JSON: {{"tweets":["The tweet"]}}
{FORMATTING_RULES}
Under 280 chars. Strong hook. Engagement driver at end."""
    try:
        result, _ = _call_claude(prompt, max_tokens=2048)
        return result.get("tweets", []) if result else [], None
    except Exception:
        return [], None


# -- LinkedIn pipeline (Ideas → Brief → Drafts) ---------------------------

LINKEDIN_VOICE_RULES = """LINKEDIN WRITING RULES (non-negotiable):
- Mobile-readable: one idea per line. Short paragraphs (1-2 sentences max).
- Blank line between every thought. White space is your weapon.
- NO hashtags unless the user explicitly asks.
- NO generic motivational tone. NO "Here's what I learned". NO "Let that sink in."
- Every sentence must contain a SPECIFIC detail: a name, a number, a date, a project, a dollar amount.
- Structure is angle-first: lead with the surprising take, then prove it with your own experience.
- Tone: operator sharing real experience. Conversational, not performative. Like texting a founder friend.
- Hook: one bold sentence that challenges a common belief. Then blank line.
- Body: 3-5 short paragraphs. Each paragraph = one proof point or insight.
- Closer: one sharp line. A question or a reframe. Not a call-to-engage.
- Length: 800-1500 characters. Tight but substantive."""


def _gather_linkedin_context(db_uid):
    """Build a rich context blob from all user data sources for LinkedIn prompts.

    Pulls from: voice profile, bookmark analysis, tweet analysis.
    Structured so the AI can see what the user reads, what they've said,
    and what resonates with their audience.
    """
    voice = get_voice_context(db_uid) if db_uid else PROFILE_CONTEXT
    parts = [f"=== VOICE PROFILE ===\n{voice}"]

    bookmark_analysis = _safe_db(db_load_analysis, db_uid, "bookmarks")
    if bookmark_analysis:
        bm_parts = []
        cats = bookmark_analysis.get("categories", [])
        if cats:
            bm_parts.append(f"Reading categories: {', '.join(c['name'] for c in cats[:8])}")
        summary = bookmark_analysis.get("summary")
        if summary:
            bm_parts.append(f"Reading summary: {summary}")
        gems = bookmark_analysis.get("gems", [])
        if gems:
            bm_parts.append("Notable bookmarks: " + "; ".join(
                f"{g.get('title','')} — {g.get('reason','')}" for g in gems[:5]
            ))
        actions = bookmark_analysis.get("actions", [])
        if actions:
            action_texts = []
            for a in actions[:3]:
                action_texts.append(a.get("text", a) if isinstance(a, dict) else str(a))
            bm_parts.append("Bookmark-based actions: " + "; ".join(action_texts))
        if bm_parts:
            parts.append("=== BOOKMARK BEHAVIOR ===\n" + "\n".join(bm_parts))

    tweet_analysis = _safe_db(db_load_analysis, db_uid, "tweets")
    if tweet_analysis:
        tw_parts = []
        summary = tweet_analysis.get("summary")
        if summary:
            tw_parts.append(f"Performance summary: {summary}")
        strategy = tweet_analysis.get("strategy", {})
        if strategy.get("best_topics"):
            tw_parts.append(f"Best-performing topics: {', '.join(strategy['best_topics'])}")
        if strategy.get("avoid_topics"):
            tw_parts.append(f"Underperforming topics: {', '.join(strategy['avoid_topics'])}")
        if strategy.get("posting_advice"):
            tw_parts.append(f"Posting insight: {strategy['posting_advice']}")
        top = tweet_analysis.get("top_performers", [])
        if top:
            tw_parts.append("Top performers: " + "; ".join(
                f"{t.get('title','')} — {t.get('why','')}" for t in top[:4]
            ))
        patterns = tweet_analysis.get("patterns", [])
        if patterns:
            tw_parts.append("Patterns: " + "; ".join(
                f"{p.get('pattern','')} ({p.get('recommendation','')})" for p in patterns[:3]
            ))
        if tw_parts:
            parts.append("=== CONTENT PERFORMANCE ===\n" + "\n".join(tw_parts))

    return "\n\n".join(parts)


def generate_linkedin_ideas(username, db_uid, seed_topic=None):
    """Step 1: Generate 8 LinkedIn-native content angles.

    Returns list of dicts with keys: title, angle, core_claim, post_type,
    why_now, proof_hint, recommended_hook.
    """
    if not CLAUDE_API_KEY:
        return []
    context = _gather_linkedin_context(db_uid)

    topic_clause = ""
    if seed_topic:
        topic_clause = f"""
The user provided a seed topic: "{seed_topic}"
Generate ideas that explore DIFFERENT angles on this topic — not 8 versions of the same take.
Each idea should find a distinct entry point into the subject."""

    prompt = f"""You are the LinkedIn Ideas Engine inside MyBookmarks.

Your job is to generate exactly 8 strong LinkedIn post ideas for this specific user.

These are not final posts.
These are not generic content prompts.
These are angle-first, proof-aware ideas for a founder/operator voice.

=== USER: {username} ===
{context}
{topic_clause}

You must use:
- the user's current topic if provided
- the user's bookmark behavior and reading patterns (provided above in context)
- the user's past content performance patterns (provided above in context)
- the user's voice profile, expertise, current focus, strong opinions, and writing examples (provided above in context)

Your goal is to find ideas at the intersection of:
1. what this user actually pays attention to
2. what this user can credibly say
3. what tends to work in their content
4. what makes a good LinkedIn post

A strong idea is:
- specific
- arguable
- grounded in proof, observation, or lived work
- useful for a founder/operator audience
- good raw material for a LinkedIn post

A weak idea is:
- broad
- generic
- inspirational
- something anyone could say
- disconnected from the user's authority

Choose exactly one post_type for each idea from:
- contrarian_lesson: Challenges a popular belief with proof from experience
- founder_story: A specific moment, decision, or failure that taught something non-obvious
- operator_framework: A reusable process or mental model extracted from doing the work
- market_observation: A pattern spotted from operating/investing that others haven't named yet
- bookmark_distillation: A synthesis of what the user has been reading, turned into a thesis
- pattern_recognition: A connection between 2-3 things that reveals a larger trend
- build_in_public_update: What happened this week/month with a real project — the messy truth

Across the 8 ideas, use at least 5 different post_type values. No more than 2 of the same type.

Return valid JSON only in this shape:
{{"ideas":[{{
  "title":"5-12 word working title",
  "angle":"One sentence: the specific framing that makes this interesting",
  "core_claim":"One falsifiable sentence: the central argument",
  "post_type":"one of the 7 values above",
  "why_now":"One sentence: why this is relevant right now, not 6 months ago",
  "proof_hint":"What specific personal experience, project, or data backs this up",
  "recommended_hook":"The literal opening line of the post — bold, specific, scroll-stopping"
}}]}}

Rules:
- Return exactly 8 ideas.
- Each idea must feel meaningfully different.
- Prefer claims over subjects.
- Prefer proof-backed observations over abstract advice.
- Do not write full posts.
- Do not include hashtags or emojis.
- Keep title and recommended_hook concise.
- If a topic is too broad, sharpen it.
- If proof is weak, narrow the claim.
- The output must sound like it belongs to one specific founder/operator, not a generic creator."""

    try:
        result, _ = _call_claude(prompt, max_tokens=3072)
        return result.get("ideas", []) if result else []
    except Exception:
        return []


def generate_linkedin_brief(username, db_uid, topic, angle):
    """Step 2: Convert idea/topic into a defensible LinkedIn post brief.

    Returns dict with: topic, angle, core_claim, post_type, target_audience,
    why_this_fits_user, proof_points, bookmark_signals, voice_constraints,
    cta_mode, hook_options, risk_flags.
    """
    if not CLAUDE_API_KEY:
        return None
    context = _gather_linkedin_context(db_uid)
    prompt = f"""You are the LinkedIn Brief Builder inside MyBookmarks.

Your job is to prepare a high-quality LinkedIn post brief for a founder/operator creator.

Do NOT write the final post.
Do NOT output prose.
Return only valid JSON.

=== USER: {username} ===
{context}

=== SELECTED IDEA ===
Topic: {topic}
Angle/context: {angle}

Use:
- the selected idea or raw topic
- the user's bookmark patterns
- the user's past content performance
- the user's voice profile
- the user's expertise, current focus, strong opinions, and writing examples

Your job is to convert a rough idea into a defensible LinkedIn angle.

The brief must answer:
- what is the post really saying?
- why can this specific user credibly say it?
- what proof or lived evidence supports it?
- what post shape fits best?
- what tone constraints matter?
- what risks should be avoided?

Choose exactly one primary post type from:
- contrarian_lesson
- founder_story
- operator_framework
- market_observation
- bookmark_distillation
- pattern_recognition
- build_in_public_update

Return JSON with exactly this shape:
{{"brief":{{
  "topic":"sharpened topic — more specific than the input",
  "angle":"the specific contrarian or surprising framing",
  "core_claim":"one sentence the user can credibly defend",
  "post_type":"one of the 7 values above",
  "target_audience":"who specifically will care about this",
  "why_this_fits_user":"what gives this user the authority to say this",
  "proof_points":["3-4 concrete proof points from their experience — real projects, numbers, decisions, outcomes"],
  "bookmark_signals":["2-3 relevant patterns from what the user has been reading"],
  "voice_constraints":["3-4 specific tone/style rules for THIS post — drawn from their voice profile and don'ts"],
  "cta_mode":"none|question|conversation|soft_invite",
  "hook_options":["3 different opening lines — bold, specific, scroll-stopping, not clickbait"],
  "risk_flags":["2-3 honest risks: generic, unsupported, preachy, too broad, etc."]
}}}}

Rules:
- "angle" must be sharper than the original topic.
- "core_claim" must be one sentence the user can credibly defend.
- "why_this_fits_user" must explain the authority to say it.
- "proof_points" should be concrete, not vague.
- "voice_constraints" should reflect real voice rules and banned language patterns.
- "cta_mode" must be one of: none, question, conversation, soft_invite
- "hook_options" must be strong but not clickbait.
- "risk_flags" should be honest and useful.
- If the topic is weak, improve it.
- If evidence is weak, note that clearly.
- Do not write the post.

Return valid JSON only."""

    try:
        result, _ = _call_claude(prompt, max_tokens=3072)
        return result.get("brief") if result else None
    except Exception:
        return None


def generate_linkedin_drafts(username, db_uid, brief):
    """Step 3: Generate 3 LinkedIn post variants from a brief.

    Returns list of dicts with: label, style_note, post, scores (from critic).
    """
    if not CLAUDE_API_KEY:
        return []
    voice = get_voice_context(db_uid) if db_uid else PROFILE_CONTEXT
    brief_text = json.dumps(brief, indent=2)
    prompt = f"""You are the LinkedIn Writing Engine inside MyBookmarks.

Your job is to write 3 strong LinkedIn post variants for a founder/operator.

=== USER VOICE ===
{voice}

=== BRIEF ===
{brief_text}

Write like a credible builder.
Not a motivational writer.
Not a ghostwritten influencer caricature.
Not a corporate marketer.

The post must sound like this user could have written it from lived experience.

Core principles:
- Start strong.
- One idea per line.
- Optimize for LinkedIn mobile readability.
- Prefer tension, proof, specificity, and hard-earned insight.
- If a sentence could apply to anyone, cut it.
- Use concrete nouns and real operating language.
- Avoid fluff, filler, cliches, and fake certainty.
- Do not fabricate numbers, names, or outcomes.
- Do not use hashtags unless explicitly requested.
- Do not use emojis unless explicitly requested.

Structure:
- Hook
- Expansion or tension
- Proof or example
- Insight / lesson
- Optional closer or soft CTA

Behavior by post_type (use the brief's post_type):
- contrarian_lesson: challenge common advice and defend it with proof
- founder_story: open from a real moment, mistake, or decision and extract a lesson
- operator_framework: teach a practical framework in at most 3 to 5 steps
- market_observation: identify a pattern and explain what it means
- bookmark_distillation: turn consumed knowledge into a personal applied takeaway
- pattern_recognition: highlight a repeated behavior or failure mode
- build_in_public_update: show what changed, what shipped, and what was learned

Before writing, silently honor:
- the brief's voice_constraints
- the brief's risk_flags
- the user's banned phrases (from voice profile)
- the user's writing examples (from voice profile)

Output rules:
- Return exactly 3 variants as JSON
- Variant 1 = sharpest / most contrarian
- Variant 2 = most personal / founder voice
- Variant 3 = most useful / educational
- Each variant: 8 to 16 short lines, plain text, no markdown
- Use \\n for line breaks within post text

Return ONLY valid JSON:
{{"drafts":[
  {{"label":"Sharpest","style_note":"Most contrarian take","post":"the full post text with \\n line breaks"}},
  {{"label":"Most Personal","style_note":"Founder voice, story-driven","post":"the full post text with \\n line breaks"}},
  {{"label":"Most Useful","style_note":"Educational, framework-driven","post":"the full post text with \\n line breaks"}}
]}}"""

    try:
        result, _ = _call_claude(prompt, max_tokens=4096)
        drafts = result.get("drafts", []) if result else []
        if drafts:
            drafts = _score_linkedin_drafts(drafts, brief, voice)
        return drafts
    except Exception:
        return []


def _score_linkedin_drafts(drafts, brief, voice):
    """Internal critic: score each draft on quality dimensions, rank them."""
    if not CLAUDE_API_KEY or not drafts:
        return drafts
    drafts_text = "\n\n---\n\n".join(
        f"VARIANT {i+1} ({d.get('label','')}):\n{d.get('post','')}" for i, d in enumerate(drafts)
    )
    prompt = f"""You are a LinkedIn draft critic. Score each variant honestly.

Brief context: {json.dumps(brief, indent=2) if isinstance(brief, dict) else brief}

Voice context (abbreviated): {voice[:500]}

Drafts to evaluate:
{drafts_text}

Score each variant (1-10) on:
- hook_strength: Does the first line stop the scroll?
- specificity: Are there real details, not vague claims?
- proof_density: How much lived evidence is in the post?
- voice_match: Does it sound like this specific person?
- genericness: How generic is it? (1=very generic, 10=very specific)
- guru_tone: How much guru/motivational tone? (1=very guru, 10=no guru at all)
- mobile_readability: Short lines, good whitespace, scannable?

Return ONLY valid JSON:
{{"scores":[
  {{"variant":1,"hook_strength":0,"specificity":0,"proof_density":0,"voice_match":0,"genericness":0,"guru_tone":0,"mobile_readability":0,"total":0,"flag":"optional one-line concern or empty string"}},
  {{"variant":2,"hook_strength":0,"specificity":0,"proof_density":0,"voice_match":0,"genericness":0,"guru_tone":0,"mobile_readability":0,"total":0,"flag":""}},
  {{"variant":3,"hook_strength":0,"specificity":0,"proof_density":0,"voice_match":0,"genericness":0,"guru_tone":0,"mobile_readability":0,"total":0,"flag":""}}
]}}
Total = sum of all 7 scores (max 70). Be honest — a 6 is fine, not everything is a 9."""

    try:
        result, _ = _call_claude(prompt, max_tokens=1024)
        scores = result.get("scores", []) if result else []
        for i, draft in enumerate(drafts):
            if i < len(scores):
                draft["scores"] = scores[i]
        # Sort by total score descending but preserve original labels
        return drafts
    except Exception:
        return drafts


def build_excel(bookmarks):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bookmarks"
    headers = ["#", "Tweet", "Author", "Username", "Date", "Likes", "RTs", "URL"]
    hfill = PatternFill(start_color="1D9BF0", end_color="1D9BF0", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill, cell.font, cell.alignment = hfill, hfont, Alignment(horizontal="center")
    for i, bm in enumerate(bookmarks, 1):
        ws.append([i, bm["text"], bm["name"], f"@{bm['username']}", bm["date"], bm["likes"], bm["retweets"], bm["url"]])
        ws.cell(row=i + 1, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in zip("ABCDEFGH", [5, 75, 22, 20, 12, 8, 6, 60]):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _safe_db(fn, *args, **kwargs):
    """Wrap DB calls so failures don't crash pages."""
    try:
        return fn(*args, **kwargs)
    except Exception:
        return None


def ensure_db_uid():
    """Make sure db_user_id is set in session, recover if missing."""
    db_uid = session.get("db_user_id")
    if db_uid:
        return db_uid
    firebase_uid = session.get("firebase_uid")
    if firebase_uid:
        session["db_user_id"] = firebase_uid
        return firebase_uid
    return None


# -- routes ----------------------------------------------------------------

@app.route("/login", methods=["POST"])
def login():
    """Verify client-side Firebase Auth ID token and set Flask session."""
    data = request.json or {}
    id_token = data.get("idToken")
    if not id_token:
        return json.dumps({"status": "error", "message": "Missing ID token"}), 400
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token["uid"]
        session["firebase_uid"] = uid
        session["db_user_id"] = uid
        session["email"] = decoded_token.get("email")
        
        # Load user context from Firestore if it exists
        user_doc = db.collection("users").document(uid).get()
        if user_doc.exists:
            ud = user_doc.to_dict()
            session["username"] = ud.get("username", decoded_token.get("name", "User"))
            session["access_token"] = ud.get("access_token")
        else:
            session["username"] = decoded_token.get("name", "User")
            session["access_token"] = None
            
        return json.dumps({"status": "success", "uid": uid}), 200
    except Exception as e:
        return json.dumps({"status": "error", "message": str(e)}), 401


@app.route("/")
def index():
    """Dashboard - unified home page."""
    configured = bool(CLIENT_ID and CLIENT_SECRET)
    firebase_uid = session.get("firebase_uid")
    if not firebase_uid:
        return render_template("index.html", configured=configured, connected=False)

    db_uid = ensure_db_uid()
    
    # Check if Twitter is connected by checking for valid token
    token = get_valid_twitter_token(db_uid)
    twitter_connected = bool(token)

    # Load all cached data
    bm_data, bm_last_id, bm_fetched = (None, None, None)
    tw_data, tw_last_id, tw_fetched = (None, None, None)
    try:
        bm_data, bm_last_id, bm_fetched = db_load_cache_full("bookmarks_cache", db_uid)
        tw_data, tw_last_id, tw_fetched = db_load_cache_full("tweets_cache", db_uid)
    except Exception:
        pass

    bm_analysis = _safe_db(db_load_analysis, db_uid, "bookmarks")
    tw_analysis = _safe_db(db_load_analysis, db_uid, "tweets")
    suggestions = _safe_db(db_load_suggestions, db_uid)
    drafts = _safe_db(db_load_drafts, db_uid, "draft") or []

    return render_template("dashboard.html",
        connected=True, 
        twitter_connected=twitter_connected, 
        configured=configured,
        username=session.get("username", "User"),
        bookmarks=bm_data, bookmarks_fetched=bm_fetched,
        bookmarks_count=len(bm_data) if bm_data else 0,
        tweets=tw_data, tweets_fetched=tw_fetched,
        tweets_count=len(tw_data) if tw_data else 0,
        bm_analysis=bm_analysis, tw_analysis=tw_analysis,
        suggestions=suggestions, drafts=drafts,
    )


@app.route("/connect", methods=["POST"])
def connect():
    if not CLIENT_ID or not CLIENT_SECRET:
        return redirect("/")
    verifier, challenge = generate_pkce()
    state = secrets.token_urlsafe(16)
    session["verifier"] = verifier
    session["state"] = state
    params = {"response_type": "code", "client_id": CLIENT_ID, "redirect_uri": get_redirect_uri(),
              "scope": SCOPES, "state": state, "code_challenge": challenge, "code_challenge_method": "S256"}
    return redirect(f"{AUTH_URL}?{urllib.parse.urlencode(params)}")


@app.route("/callback")
def callback():
    code = request.args.get("code")
    state = request.args.get("state")
    if not code or state != session.get("state"):
        return render_template("index.html", configured=True, error="Authorization failed.")
    try:
        token_data = exchange_code(code, session.get("verifier", ""))
    except Exception as e:
        return render_template("index.html", configured=True, error=f"Token exchange error: {e}")
    token = token_data.get("access_token")
    refresh_token = token_data.get("refresh_token")
    expires_in = token_data.get("expires_in", 7200)
    if not token:
        return render_template("index.html", configured=True, error=f"Token error: {token_data}")
        
    session["access_token"] = token
    try:
        uid, uname, name = get_me(token)
    except Exception as e:
        return render_template("index.html", configured=True, error=f"Failed to get user: {e}")
        
    session["user_id"] = uid
    session["username"] = uname
    session["name"] = name
    
    try:
        db_uid = db_get_or_create_user(uid, uname, name, token, refresh_token)
        expires_at = datetime.now(timezone.utc).timestamp() + expires_in - 60
        db.collection("users").document(db_uid).update({
            "expires_at": expires_at
        })
        session["db_user_id"] = db_uid
    except Exception as e:
        print(f"Error mapping callback user: {e}")
        session["db_user_id"] = session.get("firebase_uid") or uid
        
    return redirect("/")


@app.route("/api/extension/sync", methods=["POST"])
def extension_sync():
    """Endpoint for Chrome Extension to upload scraped bookmarks."""
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return json.dumps({"status": "error", "message": "Missing or invalid authorization header"}), 401
    
    id_token = auth_header.split("Bearer ")[1]
    try:
        decoded_token = auth.verify_id_token(id_token)
        uid = decoded_token["uid"]
    except Exception as e:
        return json.dumps({"status": "error", "message": f"Token verification failed: {e}"}), 401
        
    data = request.json or {}
    bookmarks = data.get("bookmarks", [])
    
    if not bookmarks:
        return json.dumps({"status": "success", "count": 0, "total": 0}), 200
        
    try:
        existing_data, _, _ = db_load_cache_full("bookmarks_cache", uid)
        if existing_data:
            existing_ids = {item["id"] for item in existing_data}
            new_bookmarks = [item for item in bookmarks if item["id"] not in existing_ids]
            merged = new_bookmarks + existing_data
        else:
            merged = bookmarks
            
        latest_id = merged[0]["id"] if merged else None
        db_save_cache("bookmarks_cache", uid, merged, latest_id)
        
        return json.dumps({"status": "success", "count": len(bookmarks), "total": len(merged)}), 200
    except Exception as e:
        return json.dumps({"status": "error", "message": str(e)}), 500


@app.route("/sync", methods=["POST"])
def sync():
    """Delta sync - fetches only new bookmarks and tweets."""
    db_uid = ensure_db_uid()
    token = get_valid_twitter_token(db_uid)
    
    user_doc = db.collection("users").document(db_uid).get()
    if not user_doc.exists:
        return redirect("/")
    user_data = user_doc.to_dict()
    uid = user_data.get("x_user_id")
    
    if not token or not uid:
        return redirect("/")

    errors = []
    stats = []
    test_limit = None if is_owner() else 20

    # Delta bookmarks
    try:
        _, bm_last_id, _ = db_load_cache_full("bookmarks_cache", db_uid)
        new_bm, bm_err = fetch_bookmarks_delta(token, uid, since_id=bm_last_id, max_items=test_limit)
        if bm_err:
            errors.append(f"Bookmarks: {bm_err}")
        elif new_bm:
            latest_id = new_bm[0]["id"]
            merged = db_merge_cache("bookmarks_cache", db_uid, new_bm, latest_id)
            stats.append(f"{len(new_bm)} new bookmarks (total: {len(merged)})")
        else:
            stats.append("Bookmarks: no new data")
    except Exception as e:
        errors.append(f"Bookmarks sync error: {e}")

    # Delta tweets
    try:
        _, tw_last_id, _ = db_load_cache_full("tweets_cache", db_uid)
        new_tw, tw_err = fetch_tweets_delta(token, uid, since_id=tw_last_id, max_pages=1 if test_limit else 5, max_items=test_limit)
        if tw_err:
            errors.append(f"Tweets: {tw_err}")
        elif new_tw:
            latest_id = new_tw[0]["id"]
            merged = db_merge_cache("tweets_cache", db_uid, new_tw, latest_id)
            stats.append(f"{len(new_tw)} new tweets (total: {len(merged)})")
        else:
            stats.append("Tweets: no new data")
    except Exception as e:
        errors.append(f"Tweets sync error: {e}")

    if errors:
        session["sync_error"] = " | ".join(errors)
    else:
        session["sync_status"] = "Sync complete! " + " | ".join(stats)
    return redirect("/")


@app.route("/bookmarks")
def bookmarks_view():
    if not session.get("firebase_uid"):
        return redirect("/")
    db_uid = ensure_db_uid()
    bookmarks, _, fetched = (None, None, None)
    try:
        bookmarks, _, fetched = db_load_cache_full("bookmarks_cache", db_uid)
    except Exception:
        pass
    analysis = _safe_db(db_load_analysis, db_uid, "bookmarks")
    return render_template("bookmarks.html", connected=True, username=session.get("username", ""),
                           bookmarks=bookmarks, analysis=analysis, fetched_at=fetched)


@app.route("/bookmarks/analyze", methods=["POST"])
def bookmarks_analyze():
    if not session.get("firebase_uid"):
        return redirect("/")
    db_uid = ensure_db_uid()
    bookmarks, _, fetched = db_load_cache_full("bookmarks_cache", db_uid)
    if not bookmarks:
        return redirect("/bookmarks")
    analysis, ai_error = analyze_bookmarks(bookmarks, session.get("username", ""))
    if analysis:
        _safe_db(db_save_analysis, db_uid, "bookmarks", analysis)
    error = f"AI error: {ai_error}" if ai_error else None
    return render_template("bookmarks.html", connected=True, username=session.get("username", ""),
                           bookmarks=bookmarks, analysis=analysis, fetched_at=fetched, error=error)


@app.route("/bookmarks/download", methods=["POST"])
def bookmarks_download():
    if not session.get("firebase_uid"):
        return redirect("/")
    db_uid = ensure_db_uid()
    bookmarks, _, _ = db_load_cache_full("bookmarks_cache", db_uid)
    if not bookmarks:
        return redirect("/bookmarks")
    buf = build_excel(bookmarks)
    return send_file(buf, as_attachment=True,
                     download_name=f"bookmarks_{session.get('username', 'x')}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/tweets")
def tweets_view():
    if not session.get("firebase_uid"):
        return redirect("/")
    db_uid = ensure_db_uid()
    tweets, _, fetched = (None, None, None)
    try:
        tweets, _, fetched = db_load_cache_full("tweets_cache", db_uid)
    except Exception:
        pass
    analysis = _safe_db(db_load_analysis, db_uid, "tweets")
    return render_template("content.html", connected=True, username=session.get("username", ""),
                           tweets=tweets, analysis=analysis, fetched_at=fetched)


@app.route("/tweets/analyze", methods=["POST"])
def tweets_analyze():
    if not session.get("firebase_uid"):
        return redirect("/")
    db_uid = ensure_db_uid()
    tweets, _, fetched = db_load_cache_full("tweets_cache", db_uid)
    if not tweets:
        return redirect("/tweets")
    analysis, ai_error = analyze_tweets(tweets, session.get("username", ""))
    if analysis:
        _safe_db(db_save_analysis, db_uid, "tweets", analysis)
    error = f"AI error: {ai_error}" if ai_error else None
    return render_template("content.html", connected=True, username=session.get("username", ""),
                           tweets=tweets, analysis=analysis, fetched_at=fetched, error=error)


@app.route("/compose")
def compose():
    if not session.get("firebase_uid"):
        return redirect("/")
    db_uid = ensure_db_uid()
    suggestions = _safe_db(db_load_suggestions, db_uid)
    drafts_list = _safe_db(db_load_drafts, db_uid, "draft") or []
    idea = request.args.get("idea", "")
    format_type = request.args.get("format", "tweet")
    return render_template("compose.html", connected=True, username=session.get("username", ""),
                           suggestions=suggestions, saved_drafts=drafts_list,
                           idea=idea, format_type=format_type)


@app.route("/compose/suggestions", methods=["POST"])
def compose_suggestions():
    if not session.get("firebase_uid"):
        return redirect("/")
    db_uid = ensure_db_uid()
    suggestions = generate_smart_suggestions(session.get("username", ""), db_uid)
    if suggestions:
        _safe_db(db_save_suggestions, db_uid, suggestions)
    drafts_list = _safe_db(db_load_drafts, db_uid, "draft") or []
    return render_template("compose.html", connected=True, username=session.get("username", ""),
                           suggestions=suggestions, saved_drafts=drafts_list)


@app.route("/compose/generate", methods=["POST"])
def compose_generate():
    if not session.get("firebase_uid"):
        return redirect("/")
    idea = request.form.get("idea", "").strip()
    format_type = request.form.get("format", "tweet")
    platform = request.form.get("platform", "x")
    if not idea:
        return redirect("/compose")

    db_uid = ensure_db_uid()
    drafts, _ = generate_draft(session.get("username", ""), idea, format_type, platform="x", db_uid=db_uid)

    return render_template("compose.html", connected=True, username=session.get("username", ""),
                           drafts=drafts, idea=idea, format_type=format_type)


@app.route("/compose/save", methods=["POST"])
def compose_save():
    if not session.get("firebase_uid"):
        return redirect("/")
    db_uid = ensure_db_uid()
    tweets_json = request.form.get("tweets", "[]")
    topic = request.form.get("topic", "")
    fmt = request.form.get("format", "tweet")
    try:
        tweets = json.loads(tweets_json)
    except json.JSONDecodeError:
        tweets = []
    if not tweets:
        return redirect("/compose")
    try:
        db_save_draft(db_uid, tweets, fmt, topic)
    except Exception as e:
        return render_template("compose.html", connected=True, username=session.get("username", ""),
                               error=f"Save failed (db_uid={db_uid}): {e}")
    return redirect("/compose")


@app.route("/compose/delete/<draft_id>", methods=["POST"])
def compose_delete(draft_id):
    if not session.get("firebase_uid"):
        return redirect("/")
    db_delete_draft(draft_id, session.get("db_user_id"))
    return redirect("/compose")


@app.route("/compose/post", methods=["POST"])
def compose_post():
    db_uid = ensure_db_uid()
    token = get_valid_twitter_token(db_uid)
    if not token:
        return redirect("/")
    tweets_json = request.form.get("tweets", "[]")
    draft_id = request.form.get("draft_id")
    try:
        tweets = json.loads(tweets_json)
    except json.JSONDecodeError:
        return redirect("/compose")
    if not tweets:
        return redirect("/compose")
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    posted, reply_to = [], None
    for tweet_text in tweets:
        body = {"text": tweet_text}
        if reply_to:
            body["reply"] = {"in_reply_to_tweet_id": reply_to}
        r = req_lib.post("https://api.twitter.com/2/tweets", headers=headers, json=body)
        data = r.json()
        if "data" in data:
            reply_to = data["data"]["id"]
            posted.append(data["data"]["id"])
        else:
            return render_template("compose.html", connected=True, username=session.get("username", ""),
                                   error=f"Post failed: {data}", drafts=tweets)
    if draft_id:
        db_mark_posted(draft_id, db_uid)
    first_id = posted[0] if posted else ""
    return render_template("compose.html", connected=True, username=session.get("username", ""),
                           success=True, posted_url=f"https://twitter.com/{session.get('username', '')}/status/{first_id}",
                           posted_count=len(posted))


@app.route("/compose/schedule", methods=["POST"])
def compose_schedule():
    if not session.get("firebase_uid"):
        return redirect("/")
    db_uid = ensure_db_uid()
    tweets_json = request.form.get("tweets", "[]")
    topic = request.form.get("topic", "")
    fmt = request.form.get("format", "tweet")
    scheduled_at = request.form.get("scheduled_at", "")
    try:
        tweets = json.loads(tweets_json)
    except json.JSONDecodeError:
        return redirect("/compose")
    if not tweets or not scheduled_at:
        return redirect("/compose")
    db.collection("users").document(db_uid).collection("drafts").document().set({
        "tweets": tweets, "format": fmt, "topic": topic,
        "status": "scheduled", "scheduled_at": scheduled_at,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    return redirect("/calendar")



@app.route("/calendar")
def calendar_view():
    if not session.get("firebase_uid"):
        return redirect("/")
    db_uid = ensure_db_uid()
    # Avoid composite index requirements by querying simply and sorting/limiting in-memory
    try:
        scheduled_docs = db.collection("users").document(db_uid).collection("drafts").where("status", "==", "scheduled").get()
        scheduled_list = []
        for doc in scheduled_docs:
            d = doc.to_dict()
            d["id"] = doc.id
            scheduled_list.append(d)
        scheduled_list = sorted(scheduled_list, key=lambda x: x.get("scheduled_at", ""))
    except Exception as e:
        print(f"Error loading scheduled drafts: {e}")
        scheduled_list = []

    try:
        posted_docs = db.collection("users").document(db_uid).collection("drafts").where("status", "==", "posted").get()
        posted_list = []
        for doc in posted_docs:
            d = doc.to_dict()
            d["id"] = doc.id
            posted_list.append(d)
        posted_list = sorted(posted_list, key=lambda x: x.get("posted_at", ""), reverse=True)[:10]
    except Exception as e:
        print(f"Error loading posted drafts: {e}")
        posted_list = []

    return render_template("calendar.html", connected=True, username=session.get("username", ""),
                           scheduled=scheduled_list, posted=posted_list)


@app.route("/calendar/post-now/<draft_id>", methods=["POST"])
def calendar_post_now(draft_id):
    db_uid = ensure_db_uid()
    token = get_valid_twitter_token(db_uid)
    if not token:
        return redirect("/")
    
    draft_ref = db.collection("users").document(db_uid).collection("drafts").document(draft_id)
    doc = draft_ref.get()
    if not doc.exists:
        return redirect("/calendar")
    draft = doc.to_dict()
    tweets = draft.get("tweets", [])
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    posted, reply_to = [], None
    for tweet_text in tweets:
        body = {"text": tweet_text}
        if reply_to:
            body["reply"] = {"in_reply_to_tweet_id": reply_to}
        r = req_lib.post("https://api.twitter.com/2/tweets", headers=headers, json=body)
        data = r.json()
        if "data" in data:
            reply_to = data["data"]["id"]
            posted.append(data["data"]["id"])
        else:
            return render_template("calendar.html", connected=True, username=session.get("username", ""),
                                   error=f"Post failed: {data}", scheduled=[], posted=[])
    db_mark_posted(draft_id, db_uid)
    return redirect("/calendar")


@app.route("/calendar/reschedule/<draft_id>", methods=["POST"])
def calendar_reschedule(draft_id):
    if not session.get("firebase_uid"):
        return redirect("/")
    db_uid = ensure_db_uid()
    new_time = request.form.get("scheduled_at", "")
    if new_time:
        db.collection("users").document(db_uid).collection("drafts").document(draft_id).update({
            "scheduled_at": new_time
        })
    return redirect("/calendar")


@app.route("/calendar/edit/<draft_id>", methods=["POST"])
def calendar_edit(draft_id):
    if not session.get("firebase_uid"):
        return redirect("/")
    db_uid = ensure_db_uid()
    tweets_json = request.form.get("tweets", "[]")
    try:
        tweets = json.loads(tweets_json)
    except json.JSONDecodeError:
        return redirect("/calendar")
    if tweets:
        db.collection("users").document(db_uid).collection("drafts").document(draft_id).update({
            "tweets": tweets
        })
    return redirect("/calendar")


@app.route("/calendar/delete/<draft_id>", methods=["POST"])
def calendar_delete(draft_id):
    if not session.get("firebase_uid"):
        return redirect("/")
    db_uid = ensure_db_uid()
    db.collection("users").document(db_uid).collection("drafts").document(draft_id).delete()
    return redirect("/calendar")


@app.route("/drafts")
def drafts_view():
    if not session.get("firebase_uid"):
        return redirect("/")
    db_uid = ensure_db_uid()
    all_drafts = db_load_drafts(db_uid) or []
    draft_list = [d for d in all_drafts if d.get("status") == "draft"]
    posted_list = [d for d in all_drafts if d.get("status") == "posted"]
    return render_template("drafts.html", connected=True, username=session.get("username", ""),
                           drafts=draft_list, posted=posted_list)


@app.route("/settings")
def settings():
    if not session.get("firebase_uid"):
        return redirect("/")
    db_uid = ensure_db_uid()
    profile = db_load_profile(db_uid) or {}
    return render_template("settings.html", connected=True, username=session.get("username", ""), profile=profile)


@app.route("/settings/save", methods=["POST"])
def settings_save():
    if not session.get("firebase_uid"):
        return redirect("/")
    db_uid = ensure_db_uid()

    examples_raw = request.form.get("voice_examples", "")
    voice_examples = [line.strip() for line in examples_raw.split("\n---\n") if line.strip()]

    data = {
        "bio": request.form.get("bio", "").strip(),
        "expertise": request.form.get("expertise", "").strip(),
        "current_focus": request.form.get("current_focus", "").strip(),
        "voice_examples": voice_examples,
        "opinions": request.form.get("opinions", "").strip(),
        "donts": request.form.get("donts", "").strip(),
    }
    db_save_profile(db_uid, data)
    return render_template("settings.html", connected=True, username=session.get("username", ""),
                           profile=data, saved=True)


@app.route("/debug")
def debug():
    if not session.get("firebase_uid"):
        return redirect("/")
    info = {
        "session_firebase_uid": session.get("firebase_uid"),
        "session_username": session.get("username"),
        "session_db_user_id": session.get("db_user_id"),
        "firebase_project": firebase_admin.get_app().project_id if firebase_admin.get_app() else None
    }
    return f"<html><body><pre>{json.dumps(info, indent=2, default=str)}</pre></body></html>"


@app.route("/api/cron")
def run_cron():
    """Auto-post scheduled tweets. Called by Vercel Cron every 15 min."""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    # Query user-by-user to avoid composite index / collection group index requirements
    due_docs = []
    try:
        users = db.collection("users").get()
        for user_doc in users:
            user_id = user_doc.id
            drafts = db.collection("users").document(user_id).collection("drafts")\
                       .where("status", "==", "scheduled").get()
            for draft in drafts:
                d = draft.to_dict()
                if d.get("scheduled_at", "") <= now:
                    # Injected properties expected by downstream processing
                    class TempDoc:
                        def __init__(self, doc_ref, data):
                            self.id = doc_ref.id
                            self.reference = doc_ref
                            self._data = data
                        def to_dict(self):
                            return self._data
                    due_docs.append(TempDoc(draft.reference, d))
    except Exception as e:
        print(f"Cron user-by-user query failed: {e}. Falling back to collection group query.")
        try:
            all_scheduled = db.collection_group("drafts").where("status", "==", "scheduled").get()
            for doc in all_scheduled:
                d = doc.to_dict()
                if d.get("scheduled_at", "") <= now:
                    due_docs.append(doc)
        except Exception as ex:
            print(f"Cron collection group fallback failed: {ex}")
                 
    if not due_docs:
        return json.dumps({"status": "nothing due", "checked_at": now}), 200

    results = []
    for doc in due_docs:
        draft = doc.to_dict()
        draft_id = doc.id
        user_id = doc.reference.parent.parent.id
        tweets = draft.get("tweets", [])
        
        token = get_valid_twitter_token(user_id)
        if not token:
            results.append({"draft_id": draft_id, "status": "no token"})
            continue

        headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
        posted, reply_to, failed = [], None, False
        for tweet_text in tweets:
            body = {"text": tweet_text}
            if reply_to:
                body["reply"] = {"in_reply_to_tweet_id": reply_to}
            r = req_lib.post("https://api.twitter.com/2/tweets", headers=headers, json=body)
            data = r.json()
            if "data" in data:
                reply_to = data["data"]["id"]
                posted.append(data["data"]["id"])
            else:
                results.append({"draft_id": draft_id, "status": "failed", "error": str(data)})
                failed = True
                break

        if not failed and posted:
            doc.reference.update({"status": "posted", "posted_at": now})
            results.append({"draft_id": draft_id, "status": "posted", "ids": posted})

    return json.dumps({"status": "ok", "processed": len(due_docs), "results": results}), 200


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    app.run(debug=False, port=port)

